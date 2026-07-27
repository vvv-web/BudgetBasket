from io import BytesIO
from zipfile import ZipFile

from openpyxl import load_workbook

from app.seed import DDS_LICENSE_ID, DEPARTMENT_ID, MODULE_ALPHA_ID, MODULE_BETA_ID, REQUEST_ID
from tests.test_api import auth, make_client


def exported_request_ids(content: bytes) -> set[str]:
    sheet = load_workbook(BytesIO(content)).active
    return {row[9] for row in sheet.iter_rows(min_row=2, values_only=True) if row[9]}


def exported_purposes(content: bytes) -> set[str]:
    sheet = load_workbook(BytesIO(content)).active
    return {row[4] for row in sheet.iter_rows(min_row=2, values_only=True) if row[4]}


def test_export_can_select_a_department_modules_and_fixed_requests(tmp_path):
    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")
    repo = client.app.state.repo
    repo.create(
        "requests",
        {
            "id": REQUEST_ID,
            "economist_id": None,
            "unit_id": MODULE_ALPHA_ID,
            "sum_plan": 0,
            "sum_fact": 0,
            "status": "approved",
            "frozen": True,
            "fixed": True,
        },
    )
    beta_request_id = "40000000-0000-0000-0000-000000000099"
    repo.create(
        "requests",
        {
            "id": beta_request_id,
            "economist_id": None,
            "unit_id": MODULE_BETA_ID,
            "sum_plan": 0,
            "sum_fact": 0,
            "status": "approved",
            "frozen": False,
        },
    )
    repo.create(
        "req_items",
        {
            "id": "80000000-0000-0000-0000-000000000099",
            "request_id": REQUEST_ID,
            "dds_id": DDS_LICENSE_ID,
            "invest_id": None,
            "name": "Доход от лицензий",
            "sum_plan": 100,
            "sum_fact": 0,
            "justification": "Доход",
            "status": "approved",
            "comment": "",
            "is_income": True,
        },
    )

    department_export = client.get(
        "/requests/export/closed",
        params={"department_ids": DEPARTMENT_ID, "statuses": "approved", "fixed_only": "true"},
        headers=admin,
    )
    assert department_export.status_code == 200
    assert exported_request_ids(department_export.content) == {REQUEST_ID}

    module_export = client.get(
        "/requests/export/closed",
        params={"module_ids": MODULE_BETA_ID, "statuses": "approved"},
        headers=admin,
    )
    assert module_export.status_code == 200
    assert exported_request_ids(module_export.content) == {beta_request_id}

    income_export = client.get(
        "/requests/export/closed",
        params={"department_ids": DEPARTMENT_ID, "statuses": "approved", "export_kind": "income"},
        headers=admin,
    )
    assert income_export.status_code == 200
    assert exported_request_ids(income_export.content) == {REQUEST_ID}
    assert exported_purposes(income_export.content) == {"Доход"}


def test_zgd_export_uses_hierarchical_summary_layout(tmp_path):
    client = make_client(tmp_path)
    zgd = auth(client, "zgd", "zgd")
    repo = client.app.state.repo
    request_id = "40000000-0000-0000-0000-000000000077"
    repo.create(
        "requests",
        {
            "id": request_id,
            "economist_id": None,
            "unit_id": MODULE_ALPHA_ID,
            "sum_plan": 0,
            "sum_fact": 0,
            "status": "approved",
            "frozen": True,
        },
    )
    repo.create(
        "req_items",
        {
            "id": "80000000-0000-0000-0000-000000000077",
            "request_id": request_id,
            "dds_id": DDS_LICENSE_ID,
            "invest_id": None,
            "name": "Лицензия",
            "sum_plan": 100,
            "sum_fact": 80,
            "justification": "",
            "status": "approved_with_changes",
            "comment": "",
            "is_income": False,
        },
    )

    exported = client.get(
        "/requests/export/closed",
        params={"statuses": "approved"},
        headers=zgd,
    )

    assert exported.status_code == 200
    workbook = load_workbook(BytesIO(exported.content))
    sheet = workbook.active
    assert sheet.title == "Заявки ЗГД"
    assert [cell.value for cell in sheet[1]] == ["Подразделение / группа / статья / заявка", "Статус", "План", "Факт", "Корректировка", "Приложения"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert [row[0] for row in rows] == ["Департамент цифровых продуктов", "ЦФО цифровых продуктов", "Лицензии и подписки", "Модуль клиентского кабинета"]
    assert rows[-1][1:5] == ("Утверждена", 100, 80, -20)
    details = workbook["Детализация заявок"]
    assert [cell.value for cell in details[1]][:16] == [
        "Подразделение",
        "Группа",
        "Модуль",
        "ID заявки",
        "Статус заявки",
        "Тип",
        "Назначение",
        "Категория",
        "Статья / проект",
        "Наименование",
        "Обоснование",
        "План",
        "Факт",
        "Корректировка",
        "Статус строки",
        "Комментарий",
    ]
    assert next(details.iter_rows(min_row=2, values_only=True))[3] == request_id


def test_zgd_export_archive_includes_attachments(tmp_path):
    client = make_client(tmp_path)
    employee = auth(client, "employee", "employee")
    zgd = auth(client, "zgd", "zgd")
    request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    item = client.post(
        f"/requests/{request['id']}/items",
        json={"dds_id": DDS_LICENSE_ID, "name": "Лицензия", "sum_plan": 100, "justification": ""},
        headers=employee,
    ).json()
    uploaded = client.post(
        f"/items/{item['id']}/files",
        files={"file": ("пояснение.png", b"attachment", "image/png")},
        headers=employee,
    )
    assert uploaded.status_code == 200
    repo = client.app.state.repo
    repo.update("requests", request["id"], {"status": "approved"})
    repo.update("req_items", item["id"], {"status": "approved", "sum_fact": 100})

    exported = client.get(
        "/requests/export/closed",
        params={"statuses": "approved", "include_files": "true"},
        headers=zgd,
    )

    assert exported.status_code == 200
    with ZipFile(BytesIO(exported.content)) as archive:
        names = archive.namelist()
    assert "Заявки_ЗГД.xlsx" in names
    assert any(name.startswith("Приложения/Департамент цифровых продуктов/ЦФО цифровых продуктов/") for name in names)
