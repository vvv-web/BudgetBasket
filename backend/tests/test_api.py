from types import SimpleNamespace

import pytest
from fastapi.testclient import TestClient

from app.seed import DEPARTMENT_ID, DDS_LICENSE_ID, ECONOMIST_ID, EMPLOYEE_ID, INVEST_PLATFORM_ID, MODULE_ALPHA_ID, MODULE_BETA_ID, REQUEST_ID
from tests.in_memory_repository import InMemoryRepository


class AllowingFileGuard:
    async def validate(self, upload):
        mime = upload.content_type or "application/octet-stream"
        return SimpleNamespace(
            valid=True,
            detected_mime_type=mime,
            size_bytes=0,
            reason_code=None,
            message=None,
            warnings=[],
        )


class RecordingFileGuard(AllowingFileGuard):
    def __init__(self):
        self.calls = 0

    async def validate(self, upload):
        self.calls += 1
        return await super().validate(upload)


class RejectingFileGuard:
    async def validate(self, upload):
        return SimpleNamespace(
            valid=False,
            detected_mime_type="application/octet-stream",
            size_bytes=10,
            reason_code="MIME_MISMATCH",
            message="Тип содержимого файла не соответствует его расширению.",
            warnings=[],
        )


class UnavailableFileGuard:
    async def validate(self, upload):
        from app.services.file_guard_client import FileGuardUnavailableError

        raise FileGuardUnavailableError


def use_file_guard(client: TestClient, file_guard) -> None:
    client.app.state.file_guard_client = file_guard
    client.app.state.file_service.file_guard = file_guard
    client.app.state.excel_service.file_guard = file_guard


def make_client(tmp_path):
    storage_root = tmp_path / "storage"
    from app.config import Settings
    from app.factory import create_app

    settings = Settings(database_url=None, s3_endpoint=None)
    app = create_app(repository=InMemoryRepository(), settings=settings)
    app.state.file_service.object_storage.root = storage_root / "uploads"
    file_guard = AllowingFileGuard()
    app.state.file_guard_client = file_guard
    app.state.file_service.file_guard = file_guard
    app.state.excel_service.file_guard = file_guard
    return TestClient(app)


def auth(client: TestClient, login: str, password: str) -> dict[str, str]:
    response = client.post("/auth/login", json={"login": login, "password": password})
    assert response.status_code == 200
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def test_login_all_roles(tmp_path):
    client = make_client(tmp_path)
    assert client.post("/auth/login", json={"login": "admin", "password": "admin"}).json()["user"]["role"] == "admin"
    assert client.post("/auth/login", json={"login": "economist", "password": "economist"}).json()["user"]["role"] == "economist"
    assert client.post("/auth/login", json={"login": "employee", "password": "employee"}).json()["user"]["role"] == "employee"


def test_employee_creates_adds_and_submits_request(tmp_path):
    client = make_client(tmp_path)
    headers = auth(client, "employee", "employee")
    created = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=headers)
    assert created.status_code == 200
    request_id = created.json()["id"]
    item = client.post(f"/requests/{request_id}/dds-items", json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000}, headers=headers)
    assert item.status_code == 200
    assert item.json()["category_id"] == "20000000-0000-0000-0000-000000000001"
    assert item.json()["status"] == "on_review"
    submitted = client.post(f"/requests/{request_id}/submit", headers=headers)
    assert submitted.status_code == 200
    assert submitted.json()["status"] == "on_review"


def test_employee_cannot_create_request_for_foreign_unit(tmp_path):
    client = make_client(tmp_path)
    employee = auth(client, "employee", "employee")
    admin = auth(client, "admin", "admin")
    unit = client.post(
        "/units",
        json={"parent_id": None, "name": "Foreign module", "type": "module", "is_active": True},
        headers=admin,
    )
    assert unit.status_code == 200
    denied = client.post("/requests", json={"unit_id": unit.json()["id"]}, headers=employee)
    assert denied.status_code == 403
    denied_admin = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=admin)
    assert denied_admin.status_code == 403


def test_employee_can_delete_draft_request_only(tmp_path):
    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")
    economist = auth(client, "economist", "economist")
    headers = auth(client, "employee", "employee")
    created = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=headers)
    assert created.status_code == 200
    request_id = created.json()["id"]

    denied_admin = client.delete(f"/requests/{request_id}", headers=admin)
    assert denied_admin.status_code == 403
    denied_economist = client.delete(f"/requests/{request_id}", headers=economist)
    assert denied_economist.status_code == 403

    deleted = client.delete(f"/requests/{request_id}", headers=headers)
    assert deleted.status_code == 200
    assert client.get(f"/requests/{request_id}", headers=headers).status_code == 404

    recreated = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=headers)
    assert recreated.status_code == 200
    request_id = recreated.json()["id"]
    assert client.post(f"/requests/{request_id}/dds-items", json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000}, headers=headers).status_code == 200
    submitted = client.post(f"/requests/{request_id}/submit", headers=headers)
    assert submitted.status_code == 200
    denied = client.delete(f"/requests/{request_id}", headers=headers)
    assert denied.status_code == 400


def test_employee_can_withdraw_edit_and_cancel_request(tmp_path):
    client = make_client(tmp_path)
    employee = auth(client, "employee", "employee")

    created = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee)
    assert created.status_code == 200
    request_id = created.json()["id"]

    item = client.post(
        f"/requests/{request_id}/dds-items",
        json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000},
        headers=employee,
    )
    assert item.status_code == 200

    submitted = client.post(f"/requests/{request_id}/submit", headers=employee)
    assert submitted.status_code == 200
    assert submitted.json()["status"] == "on_review"

    withdrawn = client.post(f"/requests/{request_id}/withdraw", headers=employee)
    assert withdrawn.status_code == 200
    assert withdrawn.json()["status"] == "draft"

    edited = client.patch(f"/dds-items/{item.json()['id']}", json={"sum_plan": 1500}, headers=employee)
    assert edited.status_code == 200

    resubmitted = client.post(f"/requests/{request_id}/submit", headers=employee)
    assert resubmitted.status_code == 200
    assert resubmitted.json()["status"] == "on_review"

    cancelled = client.post(f"/requests/{request_id}/cancel", headers=employee)
    assert cancelled.status_code == 200
    assert cancelled.json()["status"] == "cancelled"


def test_budget_freeze_requires_approved_status_and_blocks_reopen_until_unfrozen(tmp_path):
    client = make_client(tmp_path)
    employee = auth(client, "employee", "employee")
    economist = auth(client, "economist", "economist")

    created = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee)
    assert created.status_code == 200
    request_id = created.json()["id"]

    item = client.post(
        f"/requests/{request_id}/dds-items",
        json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000},
        headers=employee,
    )
    assert item.status_code == 200

    denied_early = client.post(f"/requests/{request_id}/freeze-budget", headers=economist)
    assert denied_early.status_code == 400

    assert client.post(f"/requests/{request_id}/submit", headers=employee).status_code == 200
    approved = client.post(f"/requests/{request_id}/approve-all-items", headers=economist)
    assert approved.status_code == 200
    assert approved.json()["status"] == "approved"

    frozen = client.post(f"/requests/{request_id}/freeze-budget", headers=economist)
    assert frozen.status_code == 200
    assert frozen.json()["budget_frozen"] is True
    dashboard = client.get("/dashboard", headers=economist).json()
    assert dashboard["totals"]["frozen"] == 1000
    assert dashboard["totals"]["frozen_requests_count"] == 1

    denied_reopen = client.post(f"/requests/{request_id}/reopen", headers=economist)
    assert denied_reopen.status_code == 400

    denied_patch = client.patch(f"/dds-items/{item.json()['id']}", json={"sum_plan": 1500}, headers=employee)
    assert denied_patch.status_code == 400

    denied_upload = client.post(
        f"/dds-items/{item.json()['id']}/files",
        headers=employee,
        files={"file": ("kp.pdf", b"demo pdf content", "application/pdf")},
    )
    assert denied_upload.status_code == 400

    unfrozen = client.post(f"/requests/{request_id}/unfreeze-budget", headers=economist)
    assert unfrozen.status_code == 200
    assert unfrozen.json()["budget_frozen"] is False

    reopened = client.post(f"/requests/{request_id}/reopen", headers=economist)
    assert reopened.status_code == 200
    assert reopened.json()["status"] == "on_review"

    edited = client.patch(
        f"/dds-items/{item.json()['id']}",
        json={"status": "approved_with_changes", "sum_fact": 1500, "comment": "Пересмотрено"},
        headers=economist,
    )
    assert edited.status_code == 200


def test_files_allowed_only_in_draft_for_employee(tmp_path):
    client = make_client(tmp_path)
    employee = auth(client, "employee", "employee")
    economist = auth(client, "economist", "economist")
    request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    dds_item = client.post(f"/requests/{request['id']}/dds-items", json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000}, headers=employee).json()

    upload = client.post(
        f"/dds-items/{dds_item['id']}/files",
        headers=employee,
        files={"file": ("kp.pdf", b"demo pdf content", "application/pdf")},
    )
    assert upload.status_code == 200

    assert client.post(f"/requests/{request['id']}/submit", headers=employee).status_code == 200

    denied_upload = client.post(
        f"/dds-items/{dds_item['id']}/files",
        headers=employee,
        files={"file": ("kp2.pdf", b"demo pdf content", "application/pdf")},
    )
    assert denied_upload.status_code == 400

    denied_delete = client.delete(f"/dds-items/{dds_item['id']}/files/{upload.json()['id']}", headers=employee)
    assert denied_delete.status_code == 400

    denied_economist = client.delete(f"/dds-items/{dds_item['id']}/files/{upload.json()['id']}", headers=economist)
    assert denied_economist.status_code == 403


def test_rejected_file_is_not_saved_or_registered(tmp_path):
    client = make_client(tmp_path)
    employee = auth(client, "employee", "employee")
    request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    item = client.post(
        f"/requests/{request['id']}/dds-items",
        json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000},
        headers=employee,
    ).json()
    before_files = len(client.app.state.repo.load_all("files"))
    before_objects = len(client.app.state.repo.load_all("storage_objects"))
    use_file_guard(client, RejectingFileGuard())

    response = client.post(
        f"/dds-items/{item['id']}/files",
        headers=employee,
        files={"file": ("fake.pdf", b"not a pdf", "application/pdf")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Файл «fake.pdf»: Тип содержимого файла не соответствует его расширению."
    assert len(client.app.state.repo.load_all("files")) == before_files
    assert len(client.app.state.repo.load_all("storage_objects")) == before_objects


def test_unavailable_file_guard_fails_closed(tmp_path):
    client = make_client(tmp_path)
    employee = auth(client, "employee", "employee")
    request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    item = client.post(
        f"/requests/{request['id']}/invest-items",
        json={"invest_id": INVEST_PLATFORM_ID, "sum_plan": 1000},
        headers=employee,
    ).json()
    before_files = len(client.app.state.repo.load_all("files"))
    use_file_guard(client, UnavailableFileGuard())

    response = client.post(
        f"/invest-items/{item['id']}/files",
        headers=employee,
        files={"file": ("offer.pdf", b"content", "application/pdf")},
    )

    assert response.status_code == 503
    assert response.json()["detail"] == "Файл «offer.pdf»: проверка файлов временно недоступна. Повторите попытку позже."
    assert len(client.app.state.repo.load_all("files")) == before_files


def test_upload_permissions_are_checked_before_file_guard(tmp_path):
    client = make_client(tmp_path)
    employee = auth(client, "employee", "employee")
    admin = auth(client, "admin", "admin")
    request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    item = client.post(
        f"/requests/{request['id']}/dds-items",
        json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000},
        headers=employee,
    ).json()
    guard = RecordingFileGuard()
    use_file_guard(client, guard)

    response = client.post(
        f"/dds-items/{item['id']}/files",
        headers=admin,
        files={"file": ("offer.pdf", b"content", "application/pdf")},
    )

    assert response.status_code == 403
    assert guard.calls == 0


def test_economist_reviews_finalizes_and_employee_cannot_edit_closed(tmp_path):
    client = make_client(tmp_path)
    economist = auth(client, "economist", "economist")
    employee = auth(client, "employee", "employee")
    assert client.post(f"/requests/{REQUEST_ID}/start-review", headers=economist).status_code == 200
    dds_item = client.get(f"/requests/{REQUEST_ID}/dds-items", headers=economist).json()[0]
    invest_item = client.get(f"/requests/{REQUEST_ID}/invest-items", headers=economist).json()[0]
    assert client.patch(f"/dds-items/{dds_item['id']}", json={"status": "approved"}, headers=economist).status_code == 200
    assert client.patch(f"/invest-items/{invest_item['id']}", json={"status": "rejected", "sum_fact": 0, "comment": "Не подтверждено"}, headers=economist).status_code == 200
    finalized = client.post(f"/requests/{REQUEST_ID}/finalize", headers=economist)
    assert finalized.status_code == 200
    assert finalized.json()["status"] == "partially_approved"
    denied = client.patch(f"/dds-items/{dds_item['id']}", json={"sum_plan": 1}, headers=employee)
    assert denied.status_code == 400


def test_item_review_is_saved_independently_and_validates_status_amount_pair(tmp_path):
    client = make_client(tmp_path)
    economist = auth(client, "economist", "economist")
    dds_item = client.get(f"/requests/{REQUEST_ID}/dds-items", headers=economist).json()[0]
    invest_item = client.get(f"/requests/{REQUEST_ID}/invest-items", headers=economist).json()[0]

    wrong_approved = client.patch(
        f"/dds-items/{dds_item['id']}",
        json={"status": "approved", "sum_fact": 1},
        headers=economist,
    )
    assert wrong_approved.status_code == 400

    saved = client.patch(f"/dds-items/{dds_item['id']}", json={"status": "approved"}, headers=economist)
    assert saved.status_code == 200
    assert saved.json()["sum_fact"] == dds_item["sum_plan"]
    assert client.get(f"/requests/{REQUEST_ID}", headers=economist).json()["status"] == "on_review"

    unchanged_amount = client.patch(
        f"/invest-items/{invest_item['id']}",
        json={"status": "approved_with_changes", "sum_fact": invest_item["sum_plan"]},
        headers=economist,
    )
    assert unchanged_amount.status_code == 400

    wrong_rejected = client.patch(
        f"/invest-items/{invest_item['id']}",
        json={"status": "rejected", "sum_fact": 1},
        headers=economist,
    )
    assert wrong_rejected.status_code == 400


def test_economist_sees_only_requests_sent_for_review(tmp_path):
    client = make_client(tmp_path)
    economist = auth(client, "economist", "economist")
    employee = auth(client, "employee", "employee")

    draft_request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    submitted_request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    assert client.post(
        f"/requests/{submitted_request['id']}/dds-items",
        json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000},
        headers=employee,
    ).status_code == 200
    assert client.post(f"/requests/{submitted_request['id']}/submit", headers=employee).status_code == 200

    requests_list = client.get("/requests", headers=economist)
    assert requests_list.status_code == 200
    assert all(item["id"] != draft_request["id"] for item in requests_list.json())
    assert any(item["id"] == submitted_request["id"] for item in requests_list.json())
    assert client.get(f"/requests/{draft_request['id']}", headers=economist).status_code == 403
    assert client.get(f"/requests/{submitted_request['id']}", headers=economist).status_code == 200


def test_request_counterparty_contact_is_visible_to_employee_and_economist(tmp_path):
    client = make_client(tmp_path)
    economist = auth(client, "economist", "economist")
    employee = auth(client, "employee", "employee")

    employee_contact = client.get(f"/requests/{REQUEST_ID}/counterparty-contact", headers=economist)
    economist_contact = client.get(f"/requests/{REQUEST_ID}/counterparty-contact", headers=employee)

    assert employee_contact.status_code == 200
    assert employee_contact.json()["user_id"] == EMPLOYEE_ID
    assert economist_contact.status_code == 200
    assert economist_contact.json()["user_id"] == ECONOMIST_ID


def test_only_one_economist_can_be_assigned_and_assignment_can_be_removed(tmp_path):
    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")

    second = client.post(
        "/users",
        json={"login": "economist2", "password": "economist2", "role": "economist"},
        headers=admin,
    ).json()
    assigned = client.post(
        "/economist-assignments",
        json={
            "economist_id": second["id"],
            "unit_id": MODULE_ALPHA_ID,
            "assignment_type": "module",
            "is_active": True,
        },
        headers=admin,
    )
    assert assigned.status_code == 200
    active = client.get("/economist-assignments", headers=admin).json()
    module_assignments = [item for item in active if item["unit_id"] == MODULE_ALPHA_ID]
    assert [item["economist_id"] for item in module_assignments] == [second["id"]]
    assert client.get(f"/units/{MODULE_ALPHA_ID}/responsible", headers=admin).json()["user_id"] == EMPLOYEE_ID

    removed = client.patch(
        f"/economist-assignments/{second['id']}:{MODULE_ALPHA_ID}",
        headers=admin,
    )
    assert removed.status_code == 200
    assert all(
        item["unit_id"] != MODULE_ALPHA_ID
        for item in client.get("/economist-assignments", headers=admin).json()
    )


def test_dashboard_returns_budget_distribution_only_for_allowed_units(tmp_path):
    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")
    economist = auth(client, "economist", "economist")
    employee = auth(client, "employee", "employee")

    draft = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    assert client.post(
        f"/requests/{draft['id']}/dds-items",
        json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000},
        headers=employee,
    ).status_code == 200
    cancelled = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    assert client.post(
        f"/requests/{cancelled['id']}/dds-items",
        json={"dds_id": DDS_LICENSE_ID, "sum_plan": 2000},
        headers=employee,
    ).status_code == 200
    assert client.post(f"/requests/{cancelled['id']}/submit", headers=employee).status_code == 200
    assert client.post(f"/requests/{cancelled['id']}/cancel", headers=employee).status_code == 200

    admin_dashboard = client.get("/dashboard", headers=admin)
    economist_dashboard = client.get("/dashboard", headers=economist)
    focused_dashboard = client.get("/dashboard", params={"unit_id": DEPARTMENT_ID}, headers=economist)
    foreign_dashboard = client.get("/dashboard", params={"unit_id": MODULE_BETA_ID}, headers=economist)

    assert admin_dashboard.status_code == 200
    assert [unit["id"] for unit in admin_dashboard.json()["scope"]["available_units"]] == [DEPARTMENT_ID]
    assert economist_dashboard.status_code == 200
    assert [unit["id"] for unit in economist_dashboard.json()["scope"]["available_units"]] == [DEPARTMENT_ID]
    assert economist_dashboard.json()["totals"]["planned"] == 470000
    assert focused_dashboard.json()["totals"]["planned"] == 470000
    assert foreign_dashboard.json()["totals"]["planned"] == 0


def test_reopen_returns_request_to_review_for_economist_editing(tmp_path):
    client = make_client(tmp_path)
    economist = auth(client, "economist", "economist")
    dds_item = client.get(f"/requests/{REQUEST_ID}/dds-items", headers=economist).json()[0]
    invest_item = client.get(f"/requests/{REQUEST_ID}/invest-items", headers=economist).json()[0]
    client.patch(f"/dds-items/{dds_item['id']}", json={"status": "approved"}, headers=economist)
    client.patch(f"/invest-items/{invest_item['id']}", json={"status": "rejected", "sum_fact": 0, "comment": "Не подтверждено"}, headers=economist)
    client.post(f"/requests/{REQUEST_ID}/finalize", headers=economist)
    reopened = client.post(f"/requests/{REQUEST_ID}/reopen", headers=economist)
    assert reopened.status_code == 200
    assert reopened.json()["status"] == "on_review"
    edited = client.patch(
        f"/dds-items/{dds_item['id']}",
        json={"comment": "Возврат к доработке"},
        headers=economist,
    )
    assert edited.status_code == 200


def test_economist_can_approve_all_items_with_one_action(tmp_path):
    client = make_client(tmp_path)
    employee = auth(client, "employee", "employee")
    economist = auth(client, "economist", "economist")

    request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    dds_item = client.post(f"/requests/{request['id']}/dds-items", json={"dds_id": DDS_LICENSE_ID, "sum_plan": 120000}, headers=employee).json()
    invest_item = client.post(
        f"/requests/{request['id']}/invest-items",
        json={"invest_id": INVEST_PLATFORM_ID, "sum_plan": 350000},
        headers=employee,
    ).json()

    assert client.post(f"/requests/{request['id']}/submit", headers=employee).status_code == 200
    assert client.patch(
        f"/invest-items/{invest_item['id']}",
        json={"sum_fact": 300000, "comment": "Частично подтверждено"},
        headers=economist,
    ).status_code == 200

    approved = client.post(f"/requests/{request['id']}/approve-all-items", headers=economist)
    assert approved.status_code == 200
    assert approved.json()["status"] == "approved_with_changes"

    frozen = client.post(f"/requests/{request['id']}/freeze-budget", headers=economist)
    assert frozen.status_code == 200
    assert frozen.json()["budget_frozen"] is True

    dds_after = client.get(f"/requests/{request['id']}/dds-items", headers=economist).json()
    invest_after = client.get(f"/requests/{request['id']}/invest-items", headers=economist).json()
    assert dds_after[0]["status"] == "approved"
    assert dds_after[0]["sum_fact"] == 120000
    assert invest_after[0]["status"] == "approved_with_changes"


def test_direct_upload_download_for_dds_and_invest_items(tmp_path):
    client = make_client(tmp_path)
    guard = RecordingFileGuard()
    use_file_guard(client, guard)
    employee = auth(client, "employee", "employee")
    request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    dds_item = client.post(f"/requests/{request['id']}/dds-items", json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000}, headers=employee).json()
    invest_item = client.post(
        f"/requests/{request['id']}/invest-items",
        json={"invest_id": INVEST_PLATFORM_ID, "sum_plan": 2000},
        headers=employee,
    ).json()

    dds_upload = client.post(
        f"/dds-items/{dds_item['id']}/files",
        headers=employee,
        files={"file": ("kp.pdf", b"demo pdf content", "application/pdf")},
    )
    assert dds_upload.status_code == 200
    downloaded = client.get(f"/files/{dds_upload.json()['id']}/download", headers=employee)
    assert downloaded.status_code == 200
    assert downloaded.content == b"demo pdf content"

    invest_upload = client.post(
        f"/invest-items/{invest_item['id']}/files",
        headers=employee,
        files={"file": ("plan.png", b"not really png", "image/png")},
    )
    assert invest_upload.status_code == 200
    files = client.get(f"/invest-items/{invest_item['id']}/files", headers=employee)
    assert files.status_code == 200
    assert files.json()[0]["id"] == invest_upload.json()["id"]
    assert guard.calls == 2


def test_unicode_filename_download(tmp_path):
    client = make_client(tmp_path)
    employee = auth(client, "employee", "employee")
    request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    dds_item = client.post(f"/requests/{request['id']}/dds-items", json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000}, headers=employee).json()

    upload = client.post(
        f"/dds-items/{dds_item['id']}/files",
        headers=employee,
        files={"file": ("договор.pdf", b"unicode filename", "application/pdf")},
    )
    assert upload.status_code == 200
    downloaded = client.get(f"/files/{upload.json()['id']}/download", headers=employee)
    assert downloaded.status_code == 200
    assert downloaded.content == b"unicode filename"
    assert "filename*=" in downloaded.headers["content-disposition"]


def test_employee_can_delete_item_file_attachment(tmp_path):
    client = make_client(tmp_path)
    employee = auth(client, "employee", "employee")
    request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    dds_item = client.post(f"/requests/{request['id']}/dds-items", json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000}, headers=employee).json()

    upload = client.post(
        f"/dds-items/{dds_item['id']}/files",
        headers=employee,
        files={"file": ("kp.pdf", b"demo pdf content", "application/pdf")},
    )
    assert upload.status_code == 200
    file_id = upload.json()["id"]

    deleted = client.delete(f"/dds-items/{dds_item['id']}/files/{file_id}", headers=employee)
    assert deleted.status_code == 200
    files = client.get(f"/dds-items/{dds_item['id']}/files", headers=employee)
    assert files.status_code == 200
    assert files.json() == []
    assert client.get(f"/files/{file_id}/download", headers=employee).status_code == 404


def test_foreign_request_and_role_field_restrictions(tmp_path):
    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")
    employee = auth(client, "employee", "employee")
    economist = auth(client, "economist", "economist")

    created_user = client.post(
        "/users",
        json={"login": "other", "password": "other", "role": "employee", "name": "Other", "last_name": "Employee"},
        headers=admin,
    )
    assert created_user.status_code == 200
    unit = client.post(
        "/units",
        json={"parent_id": None, "name": "Other unit", "type": "department", "is_active": True},
        headers=admin,
    )
    assert unit.status_code == 200
    assert client.post(f"/units/{unit.json()['id']}/responsible", json={"user_id": created_user.json()["id"]}, headers=admin).status_code == 200
    other = auth(client, "other", "other")
    foreign = client.post("/requests", json={"unit_id": unit.json()["id"]}, headers=other)
    assert foreign.status_code == 200
    denied = client.get(f"/requests/{foreign.json()['id']}", headers=employee)
    assert denied.status_code == 403

    dds_item = client.get(f"/requests/{REQUEST_ID}/dds-items", headers=economist).json()[0]
    economist_patch = client.patch(f"/dds-items/{dds_item['id']}", json={"sum_plan": 1}, headers=economist)
    assert economist_patch.status_code == 403

    employee_request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    item = client.post(f"/requests/{employee_request['id']}/dds-items", json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000}, headers=employee).json()
    assert client.post(f"/requests/{employee_request['id']}/submit", headers=employee).status_code == 200
    employee_patch = client.patch(f"/dds-items/{item['id']}", json={"sum_fact": 1, "status": "approved"}, headers=employee)
    assert employee_patch.status_code == 400


def test_catalog_scoped_by_module_and_excel_import_export(tmp_path):
    from io import BytesIO

    from openpyxl import Workbook

    client = make_client(tmp_path)
    guard = RecordingFileGuard()
    use_file_guard(client, guard)
    admin = auth(client, "admin", "admin")
    economist = auth(client, "economist", "economist")

    scoped = client.get("/catalog/dds", params={"module_id": MODULE_ALPHA_ID, "active_only": True}, headers=admin)
    assert scoped.status_code == 200
    assert all(item["unit_id"] == "10000000-0000-0000-0000-000000000001" for item in scoped.json())

    wb = Workbook()
    ws = wb.active
    ws.append(["Название", "Категория", "Подразделение", "Активен"])
    ws.append(["Командировки", "Операционные расходы", "Департамент цифровых продуктов", "да"])
    buffer = BytesIO()
    wb.save(buffer)
    imported = client.post(
        "/catalog/dds/import",
        headers=admin,
        files={"file": ("nsi.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 200
    assert imported.json()["created"] == 1
    assert guard.calls == 1

    dds_item = client.get(f"/requests/{REQUEST_ID}/dds-items", headers=economist).json()[0]
    invest_item = client.get(f"/requests/{REQUEST_ID}/invest-items", headers=economist).json()[0]
    client.patch(f"/dds-items/{dds_item['id']}", json={"status": "approved"}, headers=economist)
    client.patch(f"/invest-items/{invest_item['id']}", json={"status": "rejected", "sum_fact": 0, "comment": "Нет"}, headers=economist)
    finalized = client.post(f"/requests/{REQUEST_ID}/finalize", headers=economist)
    assert finalized.status_code == 200

    exported = client.get(f"/requests/{REQUEST_ID}/export", headers=admin)
    assert exported.status_code == 200
    assert exported.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    from openpyxl import load_workbook as load_exported

    book = load_exported(BytesIO(exported.content))
    assert book.sheetnames[0] == "Состав"


def test_catalog_import_preview_does_not_write_and_invalid_import_is_atomic(tmp_path):
    from io import BytesIO

    from openpyxl import Workbook

    client = make_client(tmp_path)
    guard = RecordingFileGuard()
    use_file_guard(client, guard)
    admin = auth(client, "admin", "admin")

    wb = Workbook()
    ws = wb.active
    ws.append(["Название", "Категория", "Подразделение", "Активен"])
    ws.append(["Предпросмотр", "Операционные расходы", "Департамент цифровых продуктов", "да"])
    buffer = BytesIO()
    wb.save(buffer)
    preview = client.post(
        "/catalog/dds/import",
        params={"preview": True},
        headers=admin,
        files={"file": ("preview.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview.status_code == 200
    assert preview.json()["preview"] is True
    assert preview.json()["created"] == 1
    assert all(item["name"] != "Предпросмотр" for item in client.get("/catalog/dds", headers=admin).json())

    invalid_book = Workbook()
    invalid_sheet = invalid_book.active
    invalid_sheet.append(["Название", "Категория", "Подразделение", "Активен"])
    invalid_sheet.append(["Не должен сохраниться", "Операционные расходы", "Департамент цифровых продуктов", "да"])
    invalid_sheet.append(["", "Операционные расходы", "Департамент цифровых продуктов", "да"])
    invalid_buffer = BytesIO()
    invalid_book.save(invalid_buffer)
    invalid_import = client.post(
        "/catalog/dds/import",
        headers=admin,
        files={"file": ("invalid.xlsx", invalid_buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert invalid_import.status_code == 200
    assert invalid_import.json()["errors"]
    assert all(item["name"] != "Не должен сохраниться" for item in client.get("/catalog/dds", headers=admin).json())


def test_export_includes_approved_with_changes_request(tmp_path):
    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")
    economist = auth(client, "economist", "economist")
    dds_item = client.get(f"/requests/{REQUEST_ID}/dds-items", headers=economist).json()[0]
    invest_item = client.get(f"/requests/{REQUEST_ID}/invest-items", headers=economist).json()[0]

    assert client.patch(f"/dds-items/{dds_item['id']}", json={"status": "approved"}, headers=economist).status_code == 200
    assert client.patch(
        f"/invest-items/{invest_item['id']}",
        json={"status": "approved_with_changes", "sum_fact": 300000},
        headers=economist,
    ).status_code == 200
    finalized = client.post(f"/requests/{REQUEST_ID}/finalize", headers=economist)
    assert finalized.json()["status"] == "approved_with_changes"

    exported = client.get("/requests/export/closed", headers=admin)
    assert exported.status_code == 200


def test_export_excludes_rejected_and_cancelled_requests(tmp_path):
    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")
    economist = auth(client, "economist", "economist")
    employee = auth(client, "employee", "employee")
    dds_item = client.get(f"/requests/{REQUEST_ID}/dds-items", headers=economist).json()[0]
    invest_item = client.get(f"/requests/{REQUEST_ID}/invest-items", headers=economist).json()[0]

    assert client.patch(f"/dds-items/{dds_item['id']}", json={"status": "rejected", "sum_fact": 0}, headers=economist).status_code == 200
    assert client.patch(f"/invest-items/{invest_item['id']}", json={"status": "rejected", "sum_fact": 0}, headers=economist).status_code == 200
    assert client.post(f"/requests/{REQUEST_ID}/finalize", headers=economist).json()["status"] == "rejected"

    cancelled = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    assert client.post(
        f"/requests/{cancelled['id']}/dds-items",
        json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000},
        headers=employee,
    ).status_code == 200
    assert client.post(f"/requests/{cancelled['id']}/submit", headers=employee).status_code == 200
    assert client.post(f"/requests/{cancelled['id']}/cancel", headers=employee).status_code == 200

    assert client.get("/requests/export/closed", headers=admin).status_code == 404
    rejected_export = client.get(
        "/requests/export/closed",
        params={"statuses": "rejected"},
        headers=admin,
    )
    assert rejected_export.status_code == 200


def test_swagger_uses_local_assets(tmp_path):
    client = make_client(tmp_path)
    docs = client.get("/docs")
    assert docs.status_code == 200
    assert "/docs-assets/swagger-ui-bundle.js" in docs.text
    assert "cdn.jsdelivr.net" not in docs.text
    assert client.get("/docs-assets/swagger-ui-bundle.js").status_code == 200


def test_export_with_files_returns_zip_archive(tmp_path):
    from io import BytesIO
    from zipfile import ZipFile

    from openpyxl import load_workbook

    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")
    economist = auth(client, "economist", "economist")
    employee = auth(client, "employee", "employee")
    budget_request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()
    item = client.post(
        f"/requests/{budget_request['id']}/dds-items",
        json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000},
        headers=employee,
    ).json()
    assert client.post(
        f"/dds-items/{item['id']}/files",
        headers=employee,
        files={"file": ("justification.pdf", b"attached budget file", "application/pdf")},
    ).status_code == 200
    assert client.post(f"/requests/{budget_request['id']}/submit", headers=employee).status_code == 200
    assert client.post(f"/requests/{budget_request['id']}/approve-all-items", headers=economist).status_code == 200

    exported = client.get("/requests/export/closed", params={"include_files": True}, headers=admin)
    assert exported.status_code == 200
    assert exported.headers["content-type"].startswith("application/zip")
    with ZipFile(BytesIO(exported.content)) as archive:
        assert "Утверждение_бюджета.xlsx" in archive.namelist()
        attachment = next(name for name in archive.namelist() if name.endswith("justification.pdf"))
        assert attachment.startswith("Приложения/Модуль клиентского кабинета/Лицензии и подписки/")
        assert archive.read(attachment) == b"attached budget file"
        workbook = load_workbook(BytesIO(archive.read("Утверждение_бюджета.xlsx")))
        assert workbook.sheetnames == ["Состав"]
        sheet = workbook["Состав"]
        assert sheet.cell(2, 12).hyperlink.target == attachment
        assert sheet.cell(1, 1).value == "Подразделение"
        assert sheet.auto_filter.ref == f"A1:F{sheet.max_row}"


def test_catalog_duplicate_rows_are_rejected(tmp_path):
    from uuid import uuid4

    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")

    suffix = uuid4().hex[:8]
    for path in ("/catalog/dds", "/catalog/invests"):
        category_name = f"Категория {suffix}"
        item_name = f"Статья {suffix}"

        created_category = client.post(
            path,
            json={"parent_id": None, "unit_id": MODULE_ALPHA_ID, "name": category_name, "is_active": True},
            headers=admin,
        )
        assert created_category.status_code == 200

        duplicate_category = client.post(
            path,
            json={"parent_id": None, "unit_id": MODULE_ALPHA_ID, "name": category_name, "is_active": True},
            headers=admin,
        )
        assert duplicate_category.status_code == 400

        created_item = client.post(
            path,
            json={
                "parent_id": created_category.json()["id"],
                "unit_id": MODULE_ALPHA_ID,
                "name": item_name,
                "is_active": True,
            },
            headers=admin,
        )
        assert created_item.status_code == 200

        duplicate_item = client.post(
            path,
            json={
                "parent_id": created_category.json()["id"],
                "unit_id": MODULE_ALPHA_ID,
                "name": item_name,
                "is_active": True,
            },
            headers=admin,
        )
        assert duplicate_item.status_code == 400


def test_catalog_article_must_use_category_from_same_department(tmp_path):
    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")
    other_department = client.post(
        "/units",
        json={"parent_id": None, "name": "Other department", "type": "department", "is_active": True},
        headers=admin,
    ).json()
    category = client.post(
        "/catalog/dds",
        json={"parent_id": None, "unit_id": other_department["id"], "name": "Other category", "is_active": True},
        headers=admin,
    ).json()

    denied = client.post(
        "/catalog/dds",
        json={"parent_id": category["id"], "unit_id": DEPARTMENT_ID, "name": "Foreign article", "is_active": True},
        headers=admin,
    )
    assert denied.status_code == 400


def test_inactive_catalog_items_cannot_be_used_in_request_lines(tmp_path):
    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")
    employee = auth(client, "employee", "employee")
    budget_request = client.post("/requests", json={"unit_id": MODULE_ALPHA_ID}, headers=employee).json()

    assert client.patch(f"/catalog/dds/{DDS_LICENSE_ID}", json={"is_active": False}, headers=admin).status_code == 200
    denied_create = client.post(
        f"/requests/{budget_request['id']}/dds-items",
        json={"dds_id": DDS_LICENSE_ID, "sum_plan": 1000},
        headers=employee,
    )
    assert denied_create.status_code == 400
    assert denied_create.json()["detail"] == "Нельзя использовать неактивную запись НСИ в строке заявки"


def test_used_catalog_records_cannot_be_moved_or_deleted(tmp_path):
    client = make_client(tmp_path)
    admin = auth(client, "admin", "admin")
    new_category = client.post(
        "/catalog/dds",
        json={"parent_id": None, "unit_id": DEPARTMENT_ID, "name": "New category", "is_active": True},
        headers=admin,
    ).json()

    moved = client.patch(
        f"/catalog/dds/{DDS_LICENSE_ID}",
        json={"parent_id": new_category["id"]},
        headers=admin,
    )
    assert moved.status_code == 400
    assert client.delete("/catalog/dds/20000000-0000-0000-0000-000000000001", headers=admin).status_code == 400
