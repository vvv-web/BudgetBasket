from __future__ import annotations

import re
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models import APPROVED_ITEM_STATUSES, EXPORTABLE_REQUEST_STATUSES
from app.repositories.base import Repository
from app.services.common import get_required, require_role
from app.services.file_service import FileService
from app.services.file_guard_client import FileGuardClient, require_valid_file
from app.services.permission_service import PermissionService
from app.services.request_service import RequestService


HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DEPARTMENT_FILL = PatternFill("solid", fgColor="EAF1FF")
CFO_FILL = PatternFill("solid", fgColor="F5F8FE")
ARTICLE_FILL = PatternFill("solid", fgColor="FAFBFD")
MONEY_FORMAT = '#,##0.00'

REQUEST_STATUS_LABELS = {
    "draft": "Черновик",
    "on_review": "На проверке",
    "approved": "Утверждена",
    "approved_with_changes": "Утверждена с изменениями",
    "partially_approved": "Частично утверждена",
    "rejected": "Отклонена",
    "cancelled": "Отменена",
}

ITEM_STATUS_LABELS = {
    "on_review": "На рассмотрении",
    "rejected": "Отказано",
    "approved_with_changes": "Утверждено с изменениями",
    "approved": "Утверждено",
}


class ExcelService:
    def __init__(
        self,
        repo: Repository,
        permissions: PermissionService,
        requests: RequestService,
        files: FileService,
        export_dir: Path,
        file_guard: FileGuardClient,
    ):
        self.repo = repo
        self.permissions = permissions
        self.requests = requests
        self.files = files
        self.export_dir = export_dir
        self.file_guard = file_guard
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def department_id_for_unit(self, unit_id: str | None) -> str | None:
        if not unit_id:
            return None
        unit = self.repo.get_by_id("units", unit_id)
        if not unit:
            return None
        if unit.get("parent_id"):
            return unit["parent_id"]
        return unit["id"]

    def resolve_unit_id(self, *, unit_id: str | None = None, module_id: str | None = None) -> str | None:
        if module_id:
            return self.department_id_for_unit(module_id)
        if unit_id:
            return self.department_id_for_unit(unit_id)
        return None

    def filter_catalog(
        self,
        collection: str,
        *,
        unit_id: str | None = None,
        module_id: str | None = None,
        active_only: bool = False,
        query: str | None = None,
    ) -> list[dict]:
        department_id = self.resolve_unit_id(unit_id=unit_id, module_id=module_id)
        items = self.repo.load_all(collection)
        result = []
        needle = (query or "").strip().lower()
        for item in items:
            if department_id and item.get("unit_id") not in {department_id, None}:
                # Allow global items (unit_id null) for admins browsing everything;
                # for scoped lookups require matching department.
                if unit_id or module_id:
                    if item.get("unit_id") != department_id:
                        continue
            if active_only and not item.get("is_active", True):
                continue
            if needle:
                haystack = str(item.get("name", "")).lower()
                if needle not in haystack:
                    continue
            result.append(item)
        return result

    @staticmethod
    def _style_header(ws, columns: list[str]) -> None:
        for index, title in enumerate(columns, start=1):
            cell = ws.cell(1, index, title)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _autosize(ws) -> None:
        for column in ws.columns:
            width = 12
            letter = column[0].column_letter
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(value) + 2, 48))
            ws.column_dimensions[letter].width = width

    def build_import_template(self, kind: str) -> BytesIO:
        titles = {
            "dds": "Шаблон импорта статей ДДС",
            "invests": "Шаблон импорта инвест-проектов",
        }
        if kind not in titles:
            raise HTTPException(status_code=400, detail="Неизвестный тип справочника")
        leaf_label = "Статья ДДС" if kind == "dds" else "Инвест-проект"
        wb = Workbook()
        ws = wb.active
        ws.title = "НСИ"
        columns = ["Категория", "Название", "Подразделение", "Активен"]
        self._style_header(ws, columns)
        ws.append(["Операционные расходы", f"Пример: {leaf_label}", "Департамент цифровых продуктов", "да"])
        ws.append(["Операционные расходы", "Ещё одна подкатегория", "Департамент цифровых продуктов", "да"])
        ws.append(["Капитальные затраты", "Подкатегория другой категории", "Департамент цифровых продуктов", "да"])
        note = wb.create_sheet("Инструкция")
        note["A1"] = titles[kind]
        note["A2"] = "Структура НСИ: категория → подкатегория (статья ДДС или инвест-проект)."
        note["A3"] = "Обязательные поля: Категория, Название. Рекомендуется Подразделение."
        note["A4"] = "Одинаковая Категория в нескольких строках создаёт одну категорию и несколько подкатегорий."
        note["A5"] = "Подразделение должно совпадать с названием подразделения (корневого unit)."
        note["A6"] = "Активен: да/нет, true/false, 1/0."
        self._autosize(ws)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _find_unit(self, unit_name: str | None, unit_id: str | None) -> str | None:
        if unit_id:
            unit = self.repo.get_by_id("units", str(unit_id))
            if not unit:
                raise HTTPException(status_code=400, detail=f"Подразделение {unit_id} не найдено")
            return unit["id"] if not unit.get("parent_id") else unit["parent_id"]
        if unit_name:
            name = str(unit_name).strip().lower()
            units = self.repo.load_all("units")
            match = next((unit for unit in units if not unit.get("parent_id") and unit.get("name", "").strip().lower() == name), None)
            if not match:
                match = next((unit for unit in units if unit.get("name", "").strip().lower() == name), None)
            if not match:
                raise HTTPException(status_code=400, detail=f"Подразделение «{unit_name}» не найдено")
            return match["id"] if not match.get("parent_id") else match["parent_id"]
        department = next((unit for unit in self.repo.load_all("units") if unit.get("parent_id") is None), None)
        return department["id"] if department else None

    @staticmethod
    def _as_bool(value: Any, default: bool = True) -> bool:
        if value is None or value == "":
            return default
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if text in {"1", "true", "yes", "y", "да", "истина", "активен"}:
            return True
        if text in {"0", "false", "no", "n", "нет", "ложь", "неактивен"}:
            return False
        return default

    @staticmethod
    def _normalize_header(value: Any) -> str:
        text = str(value or "").strip().lower().replace(" ", "_")
        mapping = {
            "название": "name",
            "наименование": "name",
            "имя": "name",
            "подкатегория": "name",
            "статья": "name",
            "проект": "name",
            "категория": "category",
            "category": "category",
            "category_name": "category",
            "родитель": "category",
            "parent": "category",
            "подразделение": "unit_name",
            "департамент": "unit_name",
            "unit": "unit_name",
            "unit_id": "unit_id",
            "активен": "is_active",
            "is_active": "is_active",
            "name": "name",
            "unit_name": "unit_name",
        }
        return mapping.get(text, text)

    def _ensure_category(
        self,
        collection: str,
        *,
        category_name: str,
        unit_id: str | None,
        is_active: bool,
    ) -> dict:
        name_key = category_name.strip()
        match = next(
            (
                item
                for item in self.repo.load_all(collection)
                if not item.get("parent_id")
                and item.get("name", "").strip().lower() == name_key.lower()
                and item.get("unit_id") == unit_id
            ),
            None,
        )
        if match:
            return match
        return self.repo.create(
            collection,
            {
                "parent_id": None,
                "name": name_key,
                "unit_id": unit_id,
                "is_active": is_active,
            },
        )

    def _find_leaf(self, collection: str, *, name: str, parent_id: str | None, unit_id: str | None) -> dict | None:
        return next(
            (
                item
                for item in self.repo.load_all(collection)
                if item.get("name", "").strip().lower() == name.strip().lower()
                and item.get("parent_id") == parent_id
                and item.get("unit_id") == unit_id
            ),
            None,
        )

    async def import_catalog(self, user: dict, collection: str, upload: UploadFile, *, preview: bool = False) -> dict:
        require_role(user, "admin")
        filename = (upload.filename or "").lower()
        if not filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Ожидается файл Excel (.xlsx)")
        await require_valid_file(self.file_guard, upload)
        raw = await upload.read()
        if not raw:
            raise HTTPException(status_code=400, detail="Пустой файл")
        try:
            wb = load_workbook(BytesIO(raw), data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=400, detail="Не удалось прочитать Excel-файл") from exc
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="В файле нет строк")
        headers = [self._normalize_header(value) for value in rows[0]]
        if "name" not in headers:
            raise HTTPException(status_code=400, detail="В первой строке должен быть столбец «Наименование», «Название» или «Подкатегория»")

        prepared: list[dict] = []
        errors: list[str] = []

        for row_index, values in enumerate(rows[1:], start=2):
            if not values or all(cell is None or str(cell).strip() == "" for cell in values):
                continue
            row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
            name = str(row.get("name") or "").strip()
            if not name:
                errors.append(f"Строка {row_index}: пустое название подкатегории")
                continue
            category_name = str(row.get("category") or "").strip() or None
            try:
                unit_id = self._find_unit(
                    str(row.get("unit_name")).strip() if row.get("unit_name") not in (None, "") else None,
                    str(row.get("unit_id")).strip() if row.get("unit_id") not in (None, "") else None,
                )
            except HTTPException as exc:
                errors.append(f"Строка {row_index}: {exc.detail}")
                continue

            prepared.append(
                {
                    "row": row_index,
                    "name": name,
                    "category": category_name,
                    "unit_id": unit_id,
                    "unit_name": str(row.get("unit_name") or "").strip(),
                    "is_active": self._as_bool(row.get("is_active"), True),
                }
            )

        # Импорт применяется только целиком: ошибки в любой строке не должны оставлять
        # в справочнике частично загруженные данные.
        if errors:
            return {
                "preview": preview,
                "created": 0,
                "updated": 0,
                "errors": errors,
                "rows": prepared,
                "collection": collection,
            }

        if preview:
            preview_rows = []
            created = 0
            updated = 0
            catalog = self.repo.load_all(collection)
            for item in prepared:
                parent = next(
                    (
                        entry
                        for entry in catalog
                        if item["category"]
                        and not entry.get("parent_id")
                        and entry.get("unit_id") == item["unit_id"]
                        and entry.get("name", "").strip().casefold() == item["category"].casefold()
                    ),
                    None,
                )
                existing = self._find_leaf(
                    collection,
                    name=item["name"],
                    parent_id=parent["id"] if parent else None,
                    unit_id=item["unit_id"],
                )
                action = "update" if existing else "create"
                updated += int(bool(existing))
                created += int(not existing)
                preview_rows.append({**item, "action": action})
            return {
                "preview": True,
                "created": created,
                "updated": updated,
                "errors": [],
                "rows": preview_rows,
                "collection": collection,
            }

        created = 0
        updated = 0
        for item in prepared:
            parent = None
            if item["category"]:
                parent = self._ensure_category(
                    collection,
                    category_name=item["category"],
                    unit_id=item["unit_id"],
                    is_active=True,
                )
            payload = {
                "name": item["name"],
                "parent_id": parent["id"] if parent else None,
                "unit_id": item["unit_id"],
                "is_active": item["is_active"],
            }
            existing = self._find_leaf(
                collection,
                name=item["name"],
                parent_id=payload["parent_id"],
                unit_id=item["unit_id"],
            )
            if existing:
                self.repo.update(collection, existing["id"], payload)
                updated += 1
            else:
                self.repo.create(collection, payload)
                created += 1

        return {
            "preview": False,
            "created": created,
            "updated": updated,
            "errors": [],
            "rows": prepared,
            "collection": collection,
        }

    def _unit_name(self, unit_id: str | None) -> str:
        if not unit_id:
            return ""
        unit = self.repo.get_by_id("units", unit_id)
        return unit.get("name", unit_id) if unit else unit_id

    def _department_name(self, unit_id: str | None) -> str:
        current_id = unit_id
        visited: set[str] = set()
        while current_id and current_id not in visited:
            visited.add(current_id)
            unit = self.repo.get_by_id("units", current_id)
            if not unit:
                return current_id
            if not unit.get("parent_id"):
                return unit.get("name", current_id)
            current_id = unit["parent_id"]
        return ""

    def _unit_hierarchy(self, unit_id: str | None) -> list[dict]:
        hierarchy: list[dict] = []
        current_id = unit_id
        visited: set[str] = set()
        while current_id and current_id not in visited:
            visited.add(current_id)
            unit = self.repo.get_by_id("units", current_id)
            if not unit:
                break
            hierarchy.append(unit)
            current_id = unit.get("parent_id")
        return list(reversed(hierarchy))

    def _zgd_unit_groups(self, unit_id: str | None) -> tuple[str, str, str]:
        hierarchy = self._unit_hierarchy(unit_id)
        if not hierarchy:
            return "Не указано", "Не указано", self._unit_name(unit_id)
        department = hierarchy[0].get("name") or "Не указано"
        cfo = (hierarchy[-2] if len(hierarchy) >= 2 else hierarchy[-1]).get("name") or "Не указано"
        module = hierarchy[-1].get("name") or "Не указано"
        return department, cfo, module

    def _catalog_name(self, collection: str, item_id: str | None) -> str:
        if not item_id:
            return ""
        item = self.repo.get_by_id(collection, item_id)
        if not item:
            return item_id
        return item["name"]

    def _category_name(self, collection: str, item_id: str | None) -> str:
        if not item_id:
            return ""
        item = self.repo.get_by_id(collection, item_id)
        if not item:
            return ""
        parent_id = item.get("parent_id")
        if not parent_id:
            return ""
        return self._catalog_name(collection, parent_id)

    def _request_items(self, request_id: str, is_income: bool | None = None) -> list[dict]:
        rows: list[dict] = []
        for item in self.repo.load_all("req_items"):
            if item.get("request_id") != request_id or item.get("status") == "deleted":
                continue
            if is_income is not None and bool(item.get("is_income", False)) != is_income:
                continue
            is_dds = bool(item.get("dds_id"))
            catalog, field, kind = ("dds_catalog", "dds_id", "ДДС") if is_dds else ("invests_catalog", "invest_id", "Инвест")
            rows.append(
                {
                    "kind": kind,
                    "purpose": "Доход" if item.get("is_income", False) else "Расход",
                    "item_id": item["id"],
                    "article": self._catalog_name(catalog, item.get(field)),
                    "category": self._category_name(catalog, item.get(field)),
                    "sum_plan": float(item.get("sum_plan") or 0),
                    "sum_fact": item.get("sum_fact"),
                    "status_code": item.get("status"),
                    "status": ITEM_STATUS_LABELS.get(item.get("status"), item.get("status") or ""),
                    "comment": item.get("comment") or "",
                    "name": item.get("name") or "",
                    "justification": item.get("justification") or "",
                }
            )
        return rows

    CLOSED_STATUSES = {status.value for status in EXPORTABLE_REQUEST_STATUSES} | {"rejected"}
    DEFAULT_EXPORT_STATUSES = {status.value for status in EXPORTABLE_REQUEST_STATUSES}

    def export_closed_request(self, user: dict, request_id: str) -> Path:
        request = get_required(self.repo, "requests", request_id)
        self.permissions.require_view_request(user, request)
        if request.get("status") not in self.CLOSED_STATUSES:
            raise HTTPException(status_code=400, detail="Экспорт доступен только для закрытых заявок")
        return self._write_request_workbook([request], f"request_{request_id[:8]}.xlsx")

    def export_closed_requests(
        self,
        user: dict,
        unit_id: str | None = None,
        statuses: set[str] | None = None,
        include_files: bool = False,
        *,
        department_id: str | None = None,
        department_ids: set[str] | None = None,
        module_ids: set[str] | None = None,
        fixed_only: bool = False,
        export_kind: str = "all",
        request_ids: set[str] | None = None,
    ) -> Path:
        selected_statuses = self.DEFAULT_EXPORT_STATUSES if statuses is None else statuses
        if not selected_statuses or not selected_statuses.issubset(self.CLOSED_STATUSES):
            raise HTTPException(status_code=400, detail="Выберите допустимые статусы для экспорта")
        if export_kind not in {"all", "income", "expense"}:
            raise HTTPException(status_code=400, detail="Выберите допустимый состав экспорта")
        is_income = {"income": True, "expense": False}.get(export_kind)

        selected_unit_ids = self._export_unit_ids(unit_id, department_id, department_ids, module_ids)
        requests = []
        for status in selected_statuses:
            for item in self.requests.list_requests(user, status=status):
                if request_ids is not None and item.get("id") not in request_ids:
                    continue
                if selected_unit_ids is not None and item.get("unit_id") not in selected_unit_ids:
                    continue
                if fixed_only and not item.get("frozen"):
                    continue
                requests.append(item)
        if is_income is not None:
            requests = [item for item in requests if self._request_items(item["id"], is_income)]
        if not requests:
            raise HTTPException(status_code=404, detail="Нет закрытых заявок для экспорта")
        attachments = self._collect_export_attachments(requests, is_income) if include_files else []
        is_zgd_export = user.get("role") == "zgd"
        filename = "Заявки_ЗГД.xlsx" if is_zgd_export else {"income": "Доходы_бюджета.xlsx", "expense": "Расходы_бюджета.xlsx"}.get(export_kind, "Утверждение_бюджета.xlsx")
        workbook = (
            self._write_zgd_grouped_workbook(requests, filename, attachments, is_income)
            if is_zgd_export
            else self._write_request_workbook(requests, filename, attachments, is_income)
        )
        if not include_files:
            return workbook
        return self._write_export_archive(user, workbook, attachments)

    def _export_unit_ids(
        self,
        unit_id: str | None,
        department_id: str | None,
        department_ids: set[str] | None,
        module_ids: set[str] | None,
    ) -> set[str] | None:
        """Return the units selected by an export scope.

        ``unit_id`` remains supported for existing API consumers. A department
        scope includes every nested module, while a module scope stays limited
        to the explicitly selected modules.
        """
        selected_departments = set(department_ids or set())
        if department_id:
            selected_departments.add(department_id)
        if selected_departments:
            units = {item["id"]: item for item in self.repo.load_all("units")}
            selected_units = set(module_ids or set())
            for selected_department_id in selected_departments:
                selected_units.update(self._department_unit_ids(selected_department_id, units))
            return selected_units
        if module_ids:
            return module_ids
        return {unit_id} if unit_id else None

    @staticmethod
    def _department_unit_ids(department_id: str, units: dict[str, dict]) -> set[str]:
        if department_id not in units:
            raise HTTPException(status_code=404, detail="Подразделение не найдено")
        selected = {department_id}
        changed = True
        while changed:
            changed = False
            for unit in units.values():
                if unit.get("parent_id") in selected and unit["id"] not in selected:
                    selected.add(unit["id"])
                    changed = True
        return selected

    # Compat aliases
    def export_fixed_request(self, user: dict, request_id: str) -> Path:
        return self.export_closed_request(user, request_id)

    def export_fixed_requests(self, user: dict, unit_id: str | None = None) -> Path:
        return self.export_closed_requests(user, unit_id)

    def _collect_export_attachments(self, requests: list[dict], is_income: bool | None = None) -> list[dict]:
        request_ids = {item["id"] for item in requests}
        requests_by_id = {item["id"]: item for item in requests}
        items = {
            item["id"]: item
            for item in self.repo.load_all("req_items")
            if item.get("request_id") in request_ids
            and item.get("status") != "deleted"
            and (is_income is None or bool(item.get("is_income", False)) == is_income)
        }
        links = self.repo.load_all("req_item_files")
        files = {item["id"]: item for item in self.repo.load_all("files")}
        catalogs = {
            "dds": {item["id"]: item for item in self.repo.load_all("dds_catalog")},
            "invest": {item["id"]: item for item in self.repo.load_all("invests_catalog")},
        }
        attachments = []
        written: set[str] = set()
        for link in links:
                item = items.get(link.get("req_item_id"))
                file = files.get(link.get("file_id"))
                if not item or not file:
                    continue
                request = requests_by_id[item["request_id"]]
                department_name, cfo_name, module_name = self._zgd_unit_groups(request.get("unit_id"))
                department_name = self._archive_name(department_name, "Подразделение")
                cfo_name = self._archive_name(cfo_name, "Группа")
                module_name = self._archive_name(module_name, "Модуль")
                catalog = catalogs["dds"] if item.get("dds_id") else catalogs["invest"]
                article = catalog.get(item.get("dds_id") or item.get("invest_id"), {})
                article_name = self._archive_name(article.get("name"), "Статья")
                original_name = self._archive_name(file["original_name"], "Файл")
                archive_path = f"Приложения/{department_name}/{cfo_name}/{module_name}/{article_name}/{original_name}"
                duplicate_index = 2
                base_path = archive_path
                while archive_path in written:
                    archive_path = f"{base_path}_{duplicate_index}"
                    duplicate_index += 1
                written.add(archive_path)
                attachments.append(
                    {
                        "file_id": file["id"],
                        "item_id": item["id"],
                        "module_name": module_name,
                        "article_name": article_name,
                        "original_name": original_name,
                        "archive_path": archive_path,
                    }
                )
        return attachments

    def _write_export_archive(self, user: dict, workbook: Path, attachments: list[dict]) -> Path:
        archive = self.export_dir / f"{workbook.stem}.zip"

        with zipfile.ZipFile(archive, "w", compression=zipfile.ZIP_DEFLATED) as bundle:
            bundle.write(workbook, arcname=workbook.name)
            for attachment in attachments:
                body, _file, _storage, _size, _content_type = self.files.download(user, attachment["file_id"])
                try:
                    content = body.read()
                finally:
                    close = getattr(body, "close", None)
                    if callable(close):
                        close()
                bundle.writestr(attachment["archive_path"], content)
        return archive

    def _write_zgd_grouped_workbook(
        self,
        requests: list[dict],
        filename: str,
        attachments: list[dict],
        is_income: bool | None,
    ) -> Path:
        attachments_by_item: dict[str, list[dict]] = {}
        for attachment in attachments:
            attachments_by_item.setdefault(attachment["item_id"], []).append(attachment)

        departments: dict[str, dict] = {}
        for request in requests:
            department_name, cfo_name, module_name = self._zgd_unit_groups(request.get("unit_id"))
            request_status = REQUEST_STATUS_LABELS.get(request.get("status"), request.get("status") or "")
            for item in self._request_items(request["id"], is_income):
                article_key = f'{item["kind"]}\u0000{item["article"]}'
                planned = float(item["sum_plan"] or 0)
                approved = float(item["sum_fact"] or 0) if item["status_code"] in APPROVED_ITEM_STATUSES else 0.0

                department = departments.setdefault(
                    department_name,
                    {"planned": 0.0, "approved": 0.0, "cfos": {}},
                )
                cfo = department["cfos"].setdefault(
                    cfo_name,
                    {"planned": 0.0, "approved": 0.0, "articles": {}},
                )
                article = cfo["articles"].setdefault(
                    article_key,
                    {"name": item["article"], "planned": 0.0, "approved": 0.0, "requests": {}},
                )
                request_row = article["requests"].setdefault(
                    request["id"],
                    {
                        "module": module_name,
                        "status": request_status,
                        "planned": 0.0,
                        "approved": 0.0,
                        "attachments": [],
                    },
                )
                item_attachments = attachments_by_item.get(item["item_id"], [])
                request_row["attachments"].extend(item_attachments)
                request_row["planned"] += planned
                request_row["approved"] += approved
                article["planned"] += planned
                article["approved"] += approved
                cfo["planned"] += planned
                cfo["approved"] += approved
                department["planned"] += planned
                department["approved"] += approved

        wb = Workbook()
        ws = wb.active
        ws.title = "Заявки ЗГД"
        self._style_header(ws, ["Подразделение / группа / статья / заявка", "Статус", "План", "Факт", "Корректировка", "Приложения"])

        def append_row(
            name: str,
            status: str,
            planned: float,
            approved: float,
            *,
            fill: PatternFill | None = None,
            bold: bool = False,
            indent: int = 0,
            attachments_for_row: list[dict] | None = None,
        ) -> None:
            attachment_names = []
            seen_attachments: set[str] = set()
            for attachment in attachments_for_row or []:
                if attachment["archive_path"] in seen_attachments:
                    continue
                seen_attachments.add(attachment["archive_path"])
                attachment_names.append(attachment["original_name"])
            ws.append([name, status, planned, approved, approved - planned, "\n".join(attachment_names)])
            row_number = ws.max_row
            ws.cell(row_number, 1).alignment = Alignment(indent=indent, vertical="center")
            ws.cell(row_number, 6).alignment = Alignment(wrap_text=True, vertical="center")
            for column in (3, 4, 5):
                ws.cell(row_number, column).number_format = MONEY_FORMAT
            if fill:
                for cell in ws[row_number]:
                    cell.fill = fill
                    if bold:
                        cell.font = Font(bold=True)

        for department_name in sorted(departments, key=str.casefold):
            department = departments[department_name]
            append_row(department_name, "", department["planned"], department["approved"], fill=DEPARTMENT_FILL, bold=True)
            for cfo_name in sorted(department["cfos"], key=str.casefold):
                cfo = department["cfos"][cfo_name]
                append_row(cfo_name, "", cfo["planned"], cfo["approved"], fill=CFO_FILL, bold=True, indent=1)
                articles = cfo["articles"]
                for article_key in sorted(articles, key=lambda value: articles[value]["name"].casefold()):
                    article = articles[article_key]
                    append_row(article["name"], "", article["planned"], article["approved"], fill=ARTICLE_FILL, bold=True, indent=2)
                    for request_id, request_row in sorted(article["requests"].items(), key=lambda value: value[1]["module"].casefold()):
                        append_row(
                            request_row["module"],
                            request_row["status"],
                            request_row["planned"],
                            request_row["approved"],
                            indent=3,
                            attachments_for_row=request_row["attachments"],
                        )

        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 30
        self._autosize(ws)
        ws.auto_filter.ref = f"A1:F{ws.max_row}"
        ws.freeze_panes = "A2"

        details = wb.create_sheet("Детализация заявок")
        max_attachments = max((len(items) for items in attachments_by_item.values()), default=0)
        attachment_headers = [f"Приложение {index}" for index in range(1, max_attachments + 1)]
        self._style_header(
            details,
            [
                "Подразделение",
                "Группа",
                "Модуль",
                "ID заявки",
                "Статус заявки",
                "Тип",
                "Назначение",
                "Категория",
                "Статья / проект",
                "Наименование",
                "Обоснование",
                "План",
                "Факт",
                "Корректировка",
                "Статус строки",
                "Комментарий",
                *attachment_headers,
            ],
        )
        for request in requests:
            department_name, cfo_name, module_name = self._zgd_unit_groups(request.get("unit_id"))
            request_status = REQUEST_STATUS_LABELS.get(request.get("status"), request.get("status") or "")
            items = self._request_items(request["id"], is_income)
            if not items:
                details.append(
                    [
                        department_name,
                        cfo_name,
                        module_name,
                        request["id"],
                        request_status,
                        "",
                        "",
                        "",
                        "Строки отсутствуют",
                        "",
                        "",
                        0,
                        0,
                        0,
                        "",
                        "",
                        *([""] * max_attachments),
                    ],
                )
                continue
            for item in items:
                approved = float(item["sum_fact"] or 0) if item["status_code"] in APPROVED_ITEM_STATUSES else 0.0
                row_attachments = attachments_by_item.get(item["item_id"], [])
                details.append(
                    [
                        department_name,
                        cfo_name,
                        module_name,
                        request["id"],
                        request_status,
                        item["kind"],
                        item["purpose"],
                        item["category"],
                        item["article"],
                        item["name"],
                        item["justification"],
                        item["sum_plan"],
                        approved,
                        approved - float(item["sum_plan"] or 0),
                        item["status"],
                        item["comment"],
                        *[attachment["original_name"] for attachment in row_attachments],
                        *([""] * (max_attachments - len(row_attachments))),
                    ],
                )
                for index, attachment in enumerate(row_attachments, start=17):
                    cell = details.cell(details.max_row, index)
                    cell.hyperlink = attachment["archive_path"]
                    cell.style = "Hyperlink"
        for column in (12, 13, 14):
            for row in range(2, details.max_row + 1):
                details.cell(row, column).number_format = MONEY_FORMAT
        self._autosize(details)
        details.auto_filter.ref = f"A1:P{details.max_row}"
        details.freeze_panes = "A2"

        target = self.export_dir / filename
        wb.save(target)
        return target

    @staticmethod
    def _archive_name(value: Any, fallback: str) -> str:
        name = re.sub(r'[\\/:*?"<>|\x00-\x1f]+', "_", str(value or "").strip()).strip(". ")
        return name or fallback

    def _write_request_workbook(
        self,
        requests: list[dict],
        filename: str,
        attachments: list[dict] | None = None,
        is_income: bool | None = None,
    ) -> Path:
        wb = Workbook()
        attachments_by_item: dict[str, list[dict]] = {}
        for attachment in attachments or []:
            attachments_by_item.setdefault(attachment["item_id"], []).append(attachment)
        max_attachments = max((len(items) for items in attachments_by_item.values()), default=0)
        attachment_headers = [f"Приложение {index}" for index in range(1, max_attachments + 1)]

        composition = wb.active
        composition.title = "Состав"
        self._style_header(
            composition,
            [
                "Подразделение",
                "Модуль",
                "Статус заявки",
                "Тип",
                "Назначение",
                "Категория",
                "Статья / проект",
                "Наименование",
                "Обоснование",
                "ID заявки",
                "План",
                "Факт",
                "Статус строки",
                "Комментарий",
                *attachment_headers,
            ],
        )
        for request in requests:
            module_name = self._unit_name(request.get("unit_id"))
            department_name = self._department_name(request.get("unit_id"))
            request_status = REQUEST_STATUS_LABELS.get(request.get("status"), request.get("status") or "")
            items = self._request_items(request["id"], is_income)
            if not items:
                composition.append(
                    [
                        department_name,
                        module_name,
                        request_status,
                        "",
                        "",
                        "",
                        "Строки отсутствуют",
                        "",
                        "",
                        request["id"],
                        0,
                        None,
                        "",
                        "",
                        *([""] * max_attachments),
                    ]
                )
                continue
            for item in items:
                row_attachments = attachments_by_item.get(item["item_id"], [])
                composition.append(
                    [
                        department_name,
                        module_name,
                        request_status,
                        item["kind"],
                        item["purpose"],
                        item["category"],
                        item["article"],
                        item["name"],
                        item["justification"],
                        request["id"],
                        item["sum_plan"],
                        item["sum_fact"],
                        item["status"],
                        item["comment"],
                        *[attachment["original_name"] for attachment in row_attachments],
                        *([""] * (max_attachments - len(row_attachments))),
                    ]
                )
                for index, attachment in enumerate(row_attachments, start=15):
                    file_cell = composition.cell(composition.max_row, index)
                    file_cell.hyperlink = attachment["archive_path"]
                    file_cell.style = "Hyperlink"
        for col in (11, 12):
            for row in range(2, composition.max_row + 1):
                composition.cell(row, col).number_format = MONEY_FORMAT
        self._autosize(composition)
        composition.auto_filter.ref = f"A1:G{composition.max_row}"
        composition.freeze_panes = "A2"

        target = self.export_dir / filename
        wb.save(target)
        return target
